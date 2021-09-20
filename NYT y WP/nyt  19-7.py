import requests  #pedir URL
import re  #pedir href
from bs4 import BeautifulSoup  #pedir contenido interno de tags
import datetime  #poner en excel fecha y horario
from openpyxl import Workbook,load_workbook #crear excel
import os  #saber si el archivo existe

#######
#Clase principal. pedido y carga del archivo
#Funcion __init__

filepath='nyt19-7.xlsx'

if os.path.isfile(filepath):                #Averiguar si el archivo ya existe
    wb=load_workbook(filepath)              #Cargar archivo
else:                                       #Si el archivo no existe. crearlo.
    wb=Workbook()
    sheet=wb.active
    wb.remove(wb['Sheet'])    #Borrar planilla "sheet" creada automaticamente por el excel.


#####
#Clase con datos generales. hereda de clase anterior
#  Funcion para poner fecha y horario en planilla

datetime_object = datetime.datetime.now()

DATESHEET=(datetime_object.strftime("%d"), datetime_object.strftime("%b"), datetime_object.strftime("%Y"))

sheet=wb.create_sheet(str(DATESHEET))       #Crear planilla dentro del excel, con fecha del dia como nombre

#####
#Funcion que agrega a la planilla la fecha y hora

sheet.append([" "])             #Agregar una fila en blanco por razones de estilo
sheet.append(['Fecha y Hora: ', datetime_object])       #Agregar fecha y hora

#######

#Funcion para agregar fila con información


sheet.append(['ID', 'Titulo', 'Link', 'Copete', 'Copete1', 'Copete2', 'Copete3']) #excel

#######
#Clase para pedir pagina nyt. here de clase anterior
#Funcion que pide url

url="https://www.nytimes.com/"
r=requests.get(url)
html=r.text

soup = BeautifulSoup(html, 'html.parser')

######
#Clase semi principal. hereda de clase anterior
#funcion para eliminar links de comentarios

for x in soup.find_all(class_='css-16f7co2'):  #link comentarios
    x.decompose()

#####
#funcion para ingresar al body y crear listado de articulos

body_tag = soup.find('body')
article_tag = body_tag.find_all('article')

#####

cont=1

######
#Clase articulo. hereda de clase anterior. clase principal que es llamada por el objeto
#Funcion Init contine el for 
for article in article_tag:
    
    ID=cont

    #####
    #Funcion titulos
    title_principal = article.find(class_='css-1cmu9py e1voiwgp0') #titulos principal
    title_opinion = article.find(class_='css-z9cw67 e1voiwgp0') # titulos de opinion
    title_pr_opinion = article.find(class_='css-1yxu27x e1voiwgp0') #titulos principales opinion

    title_otros = article.find('h2') #titulos principal otros

    ######

    if title_otros:
        title_interno = title_otros.find('span')

    ######
    
    if title_principal:
        Titulo=title_principal.contents[0]
    elif title_opinion:
        Titulo=title_opinion.contents[0]
    elif title_pr_opinion:
        Titulo=title_pr_opinion.contents[0]
    elif title_interno:
        Titulo=title_interno.contents[0]
    else:
        Titulo=None
        
    ### FIN TITULOS

    #Funcion Copete

    copete_general = article.find(class_='css-1pfq5u e1lfvv700') #copetes general
    copete_principal = article.find(class_='css-ip5ca7 e1lfvv701') #copete principal

    #######


    if copete_general:
        COPETE=copete_general.contents[0]
        COPETE2=""
        COPETE3=""
        COPETE4="" 
    elif copete_principal:
        copete_lista = copete_principal.find_all('li')
        for c in copete_lista:                            
            if not COPETE:
                COPETE=c.contents[0]
            elif not COPETE2:
                COPETE2=c.contents[0]
            elif not COPETE3:
                COPETE3=c.contents[0]
            elif not COPETE4:
                COPETE4=c.contents[0]
    else:
        COPETE=None
        COPETE2=None
        COPETE3=None
        COPETE4=None   

    ### COPETE FIN

    #Crear funcion externa para el sector de "patter" de links??? que es llamada cada vez.

    #Funcion links

    link_generales = article.find(class_='css-6p6lnl')  #columna principal
    link_otros = article.find(class_='css-qvz0vj eqveam61') #varios
    link_pr_editor = article.find(class_='css-1xaqcky') #editor's pick
    link_varios = article.find(class_='css-7douaa eqveam60') #varios
    link_extra = article.find(class_='css-1qj0wac eqveam61') #articulo extra


    ######
    
    if link_generales:
        link_basico = link_generales.find('a')
        link_href = link_basico.get('href')
        #que pattern sea una funcion que es llamada cada vez??
        pattern= re.search("http", link_href)
        if not pattern:
            link_adherido = 'https://www.nytimes.com' + link_href
            Link=link_adherido
        else:
            Link=link_href
    elif link_otros:
        link_basico = link_otros.find('a')
        link_href = link_basico.get('href')
        pattern= re.search("http", link_href)
        if not pattern:
            link_adherido = 'https://www.nytimes.com' + link_href
            Link=link_adherido
        else:
            Link=link_href
    elif link_pr_editor:
        link_basico= link_pr_editor.find('a')
        link_href = link_basico.get('href')
        pattern= re.search("http", link_href)
        if not pattern:
            link_adherido = 'https://www.nytimes.com' + link_href
            Link=link_adherido
        else:
            Link=link_href
    elif link_varios:
        link_basico= link_varios.find('a')
        link_href = link_basico.get('href')
        pattern= re.search("http", link_href)
        if not pattern:
            link_adherido = 'https://www.nytimes.com' + link_href
            Link=link_adherido
        else:
            Link=link_href
    elif link_extra:
        link_basico= link_extra.find('a')
        link_href = link_basico.get('href')
        pattern= re.search("http", link_href)
        if not pattern:
            link_adherido = 'https://www.nytimes.com' + link_href
            Link=link_adherido
        else:
            Link=link_href
    else:
        Link=None

    ##########


    if Titulo:
        cont = cont+1
        sheet.append([ID, Titulo, Link, COPETE, COPETE2, COPETE3, COPETE4])
    
#########
   
wb.save(filepath)