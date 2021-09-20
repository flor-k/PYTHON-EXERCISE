import requests  #pedir URL
import re  #pedir href
from bs4 import BeautifulSoup  #pedir contenido interno de tags
import datetime  #poner en excel fecha y horario
from openpyxl import Workbook,load_workbook #crear excel
import os  #saber si el archivo existe
import pandas as pd  #borrar dulicados


filepath='WASHPOST30-8.xlsx'

if os.path.isfile(filepath):                #Averiguar si el archivo ya existe
    wb=load_workbook(filepath)              #Cargar archivo
else:                                       #Si el archivo no existe. crearlo.
    wb=Workbook()
    sheet=wb.active
    wb.remove(wb['Sheet'])    #Borrar planilla "sheet" creada automaticamente por el excel.

datetime_object = datetime.datetime.now()

DATESHEET=(datetime_object.strftime("%d"), datetime_object.strftime("%b"), datetime_object.strftime("%Y"))

sheet=wb.create_sheet(str(DATESHEET))       #Crear planilla dentro del excel, con fecha del dia como nombre

sheet.append([" "])             #Agregar una fila en blanco por razones de estilo
sheet.append(['Fecha y Hora: ', datetime_object])       #Agregar fecha y hora

sheet.append(['Titulo', 'Link', 'Copete']) #excel

url="https://www.washingtonpost.com/"
r=requests.get(url)
html=r.text

soup = BeautifulSoup(html, 'html.parser')

body_tag = soup.find('body')
article_tag = body_tag.find_all('h2')
copete_tag = body_tag.find(class_='blurb normal normal-style')
#Titulo_Lista=set(())

for article in article_tag:

    title_interno = article.find('a')
    Titulo=title_interno.contents[0]
    Link = title_interno.get('href')
    Copete = copete_tag.contents[0]

    #Titulo_Lista.add(Titulo)
    #Titulo_Lista.add(Link)
    if Titulo:
        sheet.append([Titulo, Link, Copete])

article_tag = body_tag.find_all(class_='headline xx-small thin-style text-align-inherit')

for article in article_tag:

    title_interno = article.find('a')
    Titulo=title_interno.contents[0]
    Link = title_interno.get('href')
    Copete = copete_tag.contents[0]

    #Titulo_Lista.add(Titulo)
    #Titulo_Lista.add(Link)
    if Titulo:
        sheet.append([Titulo, Link, Copete])


article_tag = body_tag.find_all(class_='headline xx-small normal-style text-align-inherit')

for article in article_tag:

    title_interno = article.find('a')
    Titulo=title_interno.contents[0]
    Link = title_interno.get('href')
    Copete = copete_tag.contents[0]

    #Titulo_Lista.add(Titulo)
    #Titulo_Lista.add(Link)
    if Titulo:
        sheet.append([Titulo, Link, Copete])

#print(Titulo_Lista)
#print(len(Titulo_Lista))

article_tag = body_tag.find_all(class_='headline xx-small light-style bulleted text-align-inherit')

for article in article_tag:

    title_interno = article.find('a')
    Titulo=title_interno.contents[0]
    Link = title_interno.get('href')
    Copete = copete_tag.contents[0]

    #Titulo_Lista.add(Titulo)
    #Titulo_Lista.add(Link)
    if Titulo:
        sheet.append([Titulo, Link, Copete])


wb.save(filepath)

data = pd.read_excel(filepath)
duplicate= data.drop_duplicates(subset=None, keep='first', inplace=False)
duplicate.to_excel("prueba4-8.xlsx", index=False, sheet_name = str(DATESHEET))

